import jieba, math
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

class Fws_Bayes(object):
	
	def __init__(self):

		self.Fwsfile = 'NTUSD_all.xlsx'
		self.testfile = '360test.xlsx'
		self.resultnum = 2
		self.stopwords = []
		self.Dict = {}
		self.Fws = {}
		self.P_pos = 0
		self.P_neg = 0
		self.symbols = ('…', ' ', '”', '﹑', '#', '\n', '；', '❤', '〔', '〕', '“', '｀', '•', 'இ', '´', '」', '「', '『', '【', '】', '●', '～', '^', '~', '◆', '.', '@', '』', '', '(',  ')', '"', '\t', '，',  '。', '/',  '！',  '、',  '―',  '？','＠',  '：',   '＃',  '%',   '＆',  '（',  '）',  '《',  '》',  '［',  '］',  '｛',  '｝',  '*',  ',',  '.',   '&',  '!',  '?',  ':',  ';',   '[',  ']', '⋯', '※', '😊', '．', '{',  '}', '　', '..', '-', '+', '=', '_', '<', '>')

	def FWS(self):

		f = open('stopwords.txt', 'r')
		self.stopwords = f.read().split('\n')

		jieba.load_userdict("jiebaexpand.txt")

		wb1 = load_workbook(self.Fwsfile)
		sheetnames1 = wb1.get_sheet_names()
		ws1 = wb1.get_sheet_by_name(sheetnames1[0])

		all_pos = 0
		all_neg = 0

		for i in range(ws1.max_row):
			label = ws1['H{0}'.format(i+1)].value
			comment = ws1['D{0}'.format(i+1)].value
			if comment != None:
				splitwords = jieba.cut(comment, cut_all=False)
				for ww in splitwords:
					if ww not in self.symbols:
						if ww not in self.stopwords:
							try:
								if label == 3:
									self.Dict[ww][0] += 1
									all_neg += 1
								elif label == 1 or label == 0 or label == 2:
									self.Dict[ww][1] += 1
									all_pos += 1
							except KeyError:
								if label == 3:
									self.Dict[ww] = [1 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0]
									all_neg += 1
								elif label == 0 or label == 1 or label == 2:
									self.Dict[ww] = [0 , 1 , 0 , 0 , 0 , 0 , 0 , 0 , 0]
									all_pos += 1
								#[negnum ,posnum ,negPMI ,posPMI ,selectPMI ,selectPosOrNeg ,normalization,P(word|neg),P(word|pos)]

		allnum = all_pos + all_neg
		self.P_pos = all_pos / allnum
		self.P_neg = all_neg / allnum

		for w in self.Dict.keys():
			if  self.Dict[w][0] != 0:
				self.Dict[w][2] = self.Dict[w][0] * math.log( ( (self.Dict[w][0] / allnum ) / ( ( all_neg / allnum ) * ( ( self.Dict[w][0] + self.Dict[w][1] ) / allnum ) ) ) , 2)
			if self.Dict[w][1] != 0:
				self.Dict[w][3] = self.Dict[w][1] * math.log( ( (self.Dict[w][1] / allnum ) / ( ( all_pos / allnum ) * ( ( self.Dict[w][0] + self.Dict[w][1] ) / allnum ) ) ) , 2)
			
			if self.Dict[w][2] > self.Dict[w][3]:
				self.Dict[w][4] = self.Dict[w][2]
				self.Dict[w][5] = 3
			elif self.Dict[w][2] < self.Dict[w][3]:
				self.Dict[w][4] = self.Dict[w][3]
				self.Dict[w][5] = 1

			# if self.Dict[w][2] == max( self.Dict[w][2:4] ):
			# 	self.Dict[w][4] = self.Dict[w][2]
			# 	self.Dict[w][5] = 0
			# elif self.Dict[w][3] == max( self.Dict[w][2:4] ):
			# 	self.Dict[w][4] = self.Dict[w][3]
			# 	self.Dict[w][5] = 1
			# else:
			# 	print('error')

		numlist = []

		for i in self.Dict.keys():
			numlist.append(self.Dict[i][4])

		comax = max(numlist)
		comin = min(numlist)

		# wb2 = load_workbook('empty.xlsx')
		# sheetnames2 = wb2.get_sheet_names()
		# ws2 = wb2.get_sheet_by_name(sheetnames2[0])
		# index = 1
		# temp = 0
		for w in self.Dict.keys():
			self.Dict[w][6] = ( self.Dict[w][4] - comin ) / ( comax - comin )
		# 	ws2['A{0}'.format(index)].value = w
		# 	temp = self.Dict[w][5]
		# 	ws2['B{0}'.format(index)].value = temp
		# 	temp = self.Dict[w][6]
		# 	ws2['C{0}'.format(index)].value = temp
		# 	temp = self.Dict[w][0]
		# 	ws2['D{0}'.format(index)].value = temp
		# 	temp = self.Dict[w][1]
		# 	ws2['E{0}'.format(index)].value = temp
		# 	index += 1
		# wb2.save('Fws.xlsx')

		file = open('Fws.txt','w')

		for w in self.Dict.keys():
			if self.Dict[w][6] > 0.000463822480359476:
				self.Fws[w] = self.Dict[w]

		CountHowManyKeys = 0

		for w in self.Fws.keys():
			CountHowManyKeys += 1
			file.write( w + '\n')



		for w in self.Fws.keys():
			self.Fws[w][7] = ( self.Fws[w][0] + 1 ) / ( all_neg + CountHowManyKeys ) 
			self.Fws[w][8] = ( self.Fws[w][1] + 1 ) / ( all_pos + CountHowManyKeys ) 

	def Comparsion(self):

		jieba.load_userdict("jiebaexpand.txt")

		wb1 = load_workbook(self.testfile)
		sheetnames1 = wb1.get_sheet_names()
		ws1 = wb1.get_sheet_by_name(sheetnames1[0])

		for i in range(ws1.max_row):
			eMAP_pos = 1
			eMAP_neg = 1
			string = ''
			comment = ws1['D{0}'.format(i+1)].value
			if comment != None:
				splitwords = jieba.cut(comment, cut_all=False)
				for ww in splitwords:
					if ww not in self.symbols:
						try:
								# print(self.Dict[ww][4] * self.P_neg,self.Dict[ww][3] * self.P_pos)
								eMAP_neg = eMAP_neg * ( self.Fws[ww][7] * self.P_neg)
								eMAP_pos = eMAP_pos * ( self.Fws[ww][8] * self.P_pos)
								string = string + ww + ','
								ws1['J{0}'.format(i+1)].value = string
						except KeyError:
							pass
				if eMAP_neg > eMAP_pos:
					ws1['I{0}'.format(i+1)].value = 3
				elif eMAP_pos > eMAP_neg:
					ws1['I{0}'.format(i+1)].value = 1
		wb1.save('ANS' + self.testfile)

	def Result(self):
		
		wb1 = load_workbook('ANS' + self.testfile)
		sheetnames1 = wb1.get_sheet_names()
		ws1 = wb1.get_sheet_by_name(sheetnames1[0])

		original = ''
		predict = ''
		count_ori_1 = 0
		count_ori_0 = 0
		count_pre_0 = 0
		count_pre_1 = 0
		correct_1 = 0
		correct_0 = 0

		for i in range(ws1.max_row):
			# print(type(ws1['I{0}'.format(i+1)]))
			if ws1['I{0}'.format(i+1)].value != None:
				predict = ws1['I{0}'.format(i+1)].value
				if predict == 3:
					count_pre_0 += 1
				else:
					count_pre_1 += 1
			else:
				predict = 5

			if ws1['H{0}'.format(i+1)].value != None:
				original = ws1['H{0}'.format(i+1)].value
				if original == 3 and predict !=5:
					count_ori_0 += 1
				elif (original == 0 or original == 1 or original == 2) and predict != 5:
					count_ori_1 += 1
			else:
				original = 5

			if original == 3 and predict == 3:
				correct_0 += 1
			elif (original == 0 or original == 1 or original == 2) and predict == 1:
				correct_1 += 1

			# print(original,predict)
		pos_precision = correct_1 / (count_pre_1 + 1)
		neg_precision = correct_0 / count_pre_0
		pos_recall = correct_1 / (count_ori_1)
		neg_recall =   correct_0 / count_ori_0
		pos_F1 = (2*pos_precision*pos_recall) / (pos_recall + pos_precision)
		neg_F1 = (2*neg_precision*neg_recall) / (neg_recall + neg_precision)
		all_precision = ( correct_1 + correct_0 ) / ( count_pre_0 + count_pre_1 )
		ave_precision = ( pos_precision + neg_precision ) / 2
		ave_recall = ( pos_recall + neg_recall ) / 2
		ave_F1 = ( pos_F1 + neg_F1 ) / 2
		recall = (correct_1 + correct_0) / (count_ori_0 + count_ori_1)
		F1 = (2 * all_precision * recall) / (all_precision + recall)

		
		print("pos_precision = {0} / {1} = {2}".format(correct_1,count_pre_1,pos_precision))
		print("pos_recall = {0} / {1} = {2}".format(correct_1,count_ori_1,pos_recall))
		print("pos_F1 = {0} ".format(pos_F1))
		print("neg_precision = {0} / {1} = {2}".format(correct_0,count_pre_0,neg_precision))
		print("neg_recall = {0} / {1} = {2}".format(correct_0,count_ori_0,neg_recall))
		print("neg_F1 = {0} ".format(neg_F1))
		print("ave_precision = {0}".format(ave_precision))
		print("ave_recall = {0}".format(ave_recall))
		print("ave_F1 = {0}".format(ave_F1))
		print("-----------------------------------------")
		print("precision = {0}".format(all_precision))
		print("recall = {0}".format(recall))
		print("F1 = {0}".format(F1))


		wb = load_workbook('empty.xlsx')
		sheetnames = wb.get_sheet_names()
		ws = wb.get_sheet_by_name(sheetnames[0])

		ws['A1'].value = pos_precision
		ws['A2'].value = pos_recall
		ws['A3'].value = pos_F1
		ws['A4'].value = neg_precision
		ws['A5'].value = neg_recall
		ws['A6'].value = neg_F1
		ws['A7'].value = ave_precision
		ws['A8'].value = ave_recall
		ws['A9'].value = ave_F1
		ws['A11'].value = all_precision
		ws['A12'].value = recall
		ws['A13'].value = F1

		wb.save('result_FB3_{0}.xlsx'.format(self.resultnum))


a = Fws_Bayes()
a.FWS()
# a.Comparsion()
# a.Result()