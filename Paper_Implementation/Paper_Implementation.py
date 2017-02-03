import openpyxl
import jieba
import math
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter


class Paper_Implementation(object):

	def __init__(self):
		self.Dict = {}
		self.symbols = ('…', ' ', '”', '﹑', '#', '\n', '；', '❤', '〔', '〕', '“', '｀', '•', 'இ', '´', '」', '「', '『', '【', '】', '●', '～', '^', '~', '◆', '.', '@', '』', '', '(',  ')', '"', '\t', '，',  '。', '/',  '！',  '、',  '―',
		                '？', '＠',  '：',   '＃',  '%',   '＆',  '（',  '）',  '《',  '》',  '［',  '］',  '｛',  '｝',  '*',  ',',  '.',   '&',  '!',  '?',  ':',  ';',   '[',  ']', '⋯', '※', '😊', '．', '{',  '}', '　', '..', '-', '+', '=', '_', '<', '>')
		self.stopwords = []
		self.N_num = {}

	def Data_Source(self):

		f = open('ntusd-negative.txt', 'r')
		words = f.read().split('\n')
		for w in words:
			self.Dict[w] = [0, 0]
							#[Label,num]

		f = open('ntusd-positive.txt', 'r')
		words = f.read().split('\n')
		for w in words:
			self.Dict[w] = [1, 0]

		f = open('stopwords.txt', 'r')
		self.stopwords = f.read().split('\n')

	def Data_Clean_and_Sentiment(self):

		jieba.load_userdict("jiebaexpand.txt")
		count = 1
		filenum = 6
		find = []
		wb2 = load_workbook('empty.xlsx')
		sheetnames2 = wb2.get_sheet_names()
		ws2 = wb2.get_sheet_by_name(sheetnames2[0])

		for i in range(0, 10):
			wb1 = load_workbook('{0}.xlsx'.format(i + 1))
			sheetnames1 = wb1.get_sheet_names()
			ws1 = wb1.get_sheet_by_name(sheetnames1[0])
			for j in range(ws1.max_row):
				pos_match = 0
				neg_match = 0
				comment = ws1['D{0}'.format(j + 1)].value
				if comment != None:
					splitwords = jieba.cut(comment, cut_all=False)
					for ww in splitwords:
						if ww not in self.symbols:
							if ww not in self.stopwords:
								try:
									self.Dict[ww][1] += 1
									if self.Dict[ww][0] == 0:
										neg_match += 1
										try:
											self.N_num[i+1][0] += 1
										except KeyError:
											self.N_num[i+1] = [1,0]
									else:
										pos_match += 1
										try:
											self.N_num[i+1][1] += 1
										except KeyError:
											self.N_num[i+1] = [0,1]
								except KeyError:
									pass
					temp_score = pos_match - neg_match
					find.append(temp_score)

		Max = max(find)
		Min = min(find)

		for i in range(0, 10):
			wb1 = load_workbook('{0}.xlsx'.format(i + 1))
			sheetnames1 = wb1.get_sheet_names()
			ws1 = wb1.get_sheet_by_name(sheetnames1[0])
			for j in range(ws1.max_row):
				pos_match = 0
				neg_match = 0
				string = ''
				name = ws1['C{0}'.format(j + 1)].value
				comment = ws1['D{0}'.format(j + 1)].value
				if comment != None:
					splitwords = jieba.cut(comment, cut_all=False)
					for ww in splitwords:
						if ww not in self.symbols:
							if ww not in self.stopwords:
								try:
									self.Dict[ww][1] += 1
									string = string + ww + ','
									if self.Dict[ww][0] == 0:
										neg_match += 1
										try:
											self.N_num[i+1][0] += 1
										except KeyError:
											self.N_num[i+1] = [1,0]
									else:
										pos_match += 1
										try:
											self.N_num[i+1][1] += 1
										except KeyError:
											self.N_num[i+1] = [0,1]
								except KeyError:
									pass
					ws2['C{0}'.format(count)].value = name
					ws2['J{0}'.format(count)].value = string
					ws2['D{0}'.format(count)].value = comment
					v = pos_match - neg_match
					finalscore = (-7) + ((v - Min) * (14) / (Max - Min))
					if string == '':
						ws2['I{0}'.format(count)].value = 0
					else:
						ws2['I{0}'.format(count)].value = round(finalscore)
					count += 1
		wb2.save('sentiment_{0}.xlsx'.format(filenum))

	# def mix(self):

	# 	wb2 = load_workbook('empty.xlsx')
	# 	sheetnames2 = wb2.get_sheet_names()
	# 	ws2 = wb2.get_sheet_by_name(sheetnames2[0])

	# 	for i in range(0, 10):
	# 		scores = {}
	# 		wb1 = load_workbook('sentiment_{0}.xlsx'.format(i + 1))
	# 		sheetnames1 = wb1.get_sheet_names()
	# 		ws1 = wb1.get_sheet_by_name(sheetnames1[0])
	# 		for j in range(ws1.max_row):
	# 			score = ws1['I{0}'.format(j+1)].value
	# 			if score != None:
	# 				try:
	# 					scores[score] += 1
	# 				except KeyError:
	# 					scores[score] = 1
	# 		for l in range(-7, 8):
	# 				try:
	# 					if ws2['A{0}'.format(l+9)].value != None:
	# 						ws2['A{0}'.format(l+9)].value =  + scores[l]
	# 				except KeyError:
	# 					scores[l] = 0
	# 					ws2['A{0}'.format(l+9)].value = 0
		
	# 	wb2.save('P1.xlsx')


	def statistic(self):
		wb2 = load_workbook('empty.xlsx')
		sheetnames2 = wb2.get_sheet_names()
		ws2 = wb2.get_sheet_by_name(sheetnames2[0])
		ws2['A1'].value = 'Score'
		ws2['B1'].value = '2015-08-05~15'
		ws2['C1'].value = '2015-09-05~15'
		ws2['D1'].value = '2015-10-05~15'
		ws2['E1'].value = '2015-11-05~15'
		ws2['F1'].value = '2015-12-05~15'
		ws2['G1'].value = '2016-01-05~15'

		for k in range(-7, 8):
			ws2['A{0}'.format(k+9)].value = k

		ws2['A17'].value = 'S_avg'
		ws2['A18'].value = 'S_net'
		ws2['A19'].value = 'S_ratio'
		ws2['A20'].value = 'S_pas'

		jieba.load_userdict("jiebaexpand.txt")

		for i in range(0, 6):
			N_love = 0
			N_hate = 0
			scores = {}
			wb1 = load_workbook('sentiment_{0}.xlsx'.format(i + 1))
			sheetnames1 = wb1.get_sheet_names()
			ws1 = wb1.get_sheet_by_name(sheetnames1[0])
			for j in range(ws1.max_row):
				comment = ws1['D{0}'.format(j + 1)].value
				if comment != None:
					splitwords = jieba.cut(comment, cut_all=False)
					for ww in splitwords:
						if ww not in self.symbols:
							if ww not in self.stopwords:
								try:
									self.Dict[ww][1] += 1
									if self.Dict[ww][0] == 0:
										# neg_match += 1
										try:
											self.N_num[i+1][0] += 1
										except KeyError:
											self.N_num[i+1] = [1,0]
									else:
										# pos_match += 1
										try:
											self.N_num[i+1][1] += 1
										except KeyError:
											self.N_num[i+1] = [0,1]
								except KeyError:
									pass
			for j in range(ws1.max_row):
				score = ws1['I{0}'.format(j+1)].value
				if score != None:
					try:
						scores[score] += 1
					except KeyError:
						scores[score] = 1
			for l in range(-7, 8):
				if i == 0:
					try:
						ws2['B{0}'.format(l+9)].value = scores[l]
						if l < -3:
							N_hate = N_hate + scores[l]
						elif l > 3:
							N_love = N_love + scores[l]
					except KeyError:
						scores[l] = 0
						ws2['B{0}'.format(l+9)].value = 0
				elif i == 1:
					try:
						ws2['C{0}'.format(l+9)].value = scores[l]
						if l < -3:
							N_hate = N_hate + scores[l]
						elif l > 3:
							N_love = N_love + scores[l]
					except KeyError:
						scores[l] = 0
						ws2['C{0}'.format(l+9)].value = 0
				elif i == 2:
					try:
						ws2['D{0}'.format(l+9)].value = scores[l]
						if l < -3:
							N_hate = N_hate + scores[l]
						elif l > 3:
							N_love = N_love + scores[l]
					except KeyError:
						scores[l] = 0
						ws2['D{0}'.format(l+9)].value = 0
				elif i == 3:
					try:
						ws2['E{0}'.format(l+9)].value = scores[l]
						if l < -3:
							N_hate = N_hate + scores[l]
						elif l > 3:
							N_love = N_love + scores[l]
					except KeyError:
						scores[l] = 0
						ws2['E{0}'.format(l+9)].value = 0
				elif i == 4:
					try:
						ws2['F{0}'.format(l+9)].value = scores[l]
						if l < -3:
							N_hate = N_hate + scores[l]
						elif l > 3:
							N_love = N_love + scores[l]
					except KeyError:
						scores[l] = 0
						ws2['F{0}'.format(l+9)].value = 0
				elif i == 5:
					try:
						ws2['G{0}'.format(l+9)].value = scores[l]
						if l < -3:
							N_hate = N_hate + scores[l]
						elif l > 3:
							N_love = N_love + scores[l]
					except KeyError:
						scores[l] = 0
						ws2['G{0}'.format(l+9)].value = 0


			N = ws1.max_row
			
			up = 0
			for m in range(-7,8):
					up = up + ( scores[m] * m )

			S_avg = up / N
			S_net = ( self.N_num[i+1][1] - self.N_num[i+1][0] ) / ( self.N_num[i+1][1] + self.N_num[i+1][0] )
			S_ratio = self.N_num[i+1][1] / self.N_num[i+1][0]
			S_pas = ( N_love + N_hate ) / ( self.N_num[i+1][1] + self.N_num[i+1][0] )
			
			if i == 0:
				ws2['B17'].value = S_avg
				ws2['B18'].value = S_net
				ws2['B19'].value = S_ratio
				ws2['B20'].value = S_pas
			elif i == 1:
				ws2['C17'].value = S_avg
				ws2['C18'].value = S_net
				ws2['C19'].value = S_ratio
				ws2['C20'].value = S_pas
			elif i == 2:
				ws2['D17'].value = S_avg
				ws2['D18'].value = S_net
				ws2['D19'].value = S_ratio
				ws2['D20'].value = S_pas
			elif i == 3:
				ws2['E17'].value = S_avg
				ws2['E18'].value = S_net
				ws2['E19'].value = S_ratio
				ws2['E20'].value = S_pas
			elif i == 4:
				ws2['F17'].value = S_avg
				ws2['F18'].value = S_net
				ws2['F19'].value = S_ratio
				ws2['F20'].value = S_pas
			elif i == 5:
				ws2['G17'].value = S_avg
				ws2['G18'].value = S_net
				ws2['G19'].value = S_ratio
				ws2['G20'].value = S_pas


		wb2.save('statistic_ma.xlsx')





a = Paper_Implementation()
a.Data_Source()
# a.Data_Clean_and_Sentiment()
# a.mix()
a.statistic()

