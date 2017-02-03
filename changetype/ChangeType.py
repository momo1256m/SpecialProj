import openpyxl
import jieba
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter


class changetype(object):

    # def changeforyanwen(self):

    #     symbols = ('…', ' ', '”', '﹑', '#', '；', '❤', '〔', '〕', '“', '｀', '•', 'இ', '´', '」', '「', '『', '【', '】', '●', '～', '^', '~', '◆', '.', '@', '』', '', '(',  ')', '"', '/t', '，',  '。', '/',  '！',  '、',  '―',  '？',
    #                '＠',  '：',   '＃',  '%',   '＆',  '（',  '）',  '《',  '》',  '［',  '］',  '｛',  '｝',  '*',  ',',  '.',   '&',  '!',  '?',  ':',  ';',   '[',  ']', '⋯', '※', '😊', '．', '{',  '}', '　', '..', '-', '+', '=', '_', '<', '>')
    #     jieba.load_userdict("jiebaexpand.txt")
    #     temp = ''
    #     k = 2

    #     wb2 = load_workbook('empty.xlsx')
    #     sheetnames2 = wb2.get_sheet_names()
    #     ws2 = wb2.get_sheet_by_name(sheetnames2[0])

    #     ws2['A1'].value = 'Comment'
    #     ws2['B1'].value = 'Label'

    #     for i in range(0, 4):
    #         wb1 = load_workbook('{0}.xlsx'.format(i + 1))
    #         sheetnames1 = wb1.get_sheet_names()
    #         ws1 = wb1.get_sheet_by_name(sheetnames1[0])

    #         for j in range(700):
    #             comment = ws1['D{0}'.format(j + 1)].value
    #             label = ws1['H{0}'.format(j + 1)].value
    #             index = 1
    #             if comment != None and label != None:
    #                 splitwords = jieba.cut(comment, cut_all=False)
    #                 for ww in splitwords:
    #                     if ww not in symbols:
    #                         if index == 1:
    #                        		temp = ww
    #                             index = 0
    #                         else:
    #                             temp = temp + ' ' + ww
    #                 ws2['A{0}'.format(k)].value = temp.replace('\n', '，')
    #                 ws2['B{0}'.format(k)].value = label
    #                 k += 1

    #     wb2.save('foryanwen.xlsx')

    # def changeforme(self):
    # 	symbols = ('…', ' ', '”', '﹑', '#', '\n', '；', '❤', '〔', '〕', '“', '｀', '•', 'இ', '´', '」', '「', '『', '【', '】', '●', '～', '^', '~', '◆', '.', '@', '』', '', '(',  ')', '"', '\t', '，',  '。', '/',  '！',  '、',  '―',  '？','＠',  '：',   '＃',  '%',   '＆',  '（',  '）',  '《',  '》',  '［',  '］',  '｛',  '｝',  '*',  ',',  '.',   '&',  '!',  '?',  ':',  ';',   '[',  ']', '⋯', '※', '😊', '．', '{',  '}', '　', '..', '-', '+', '=', '_', '<', '>')
    # 	jieba.load_userdict("jiebaexpand.txt")
    # 	f0 = open('neg.txt', 'w')
    # 	f1 = open('pos.txt', 'w')
    # 	for i in range(0, 4):
    # 		wb1 = load_workbook('{0}.xlsx'.format(i + 1))
    # 		sheetnames1 = wb1.get_sheet_names()
    # 		ws1 = wb1.get_sheet_by_name(sheetnames1[0])
    # 		for j in range(700):
    # 			comment = ws1['D{0}'.format(j + 1)].value
    # 			label = ws1['H{0}'.format(j + 1)].value
    # 			if comment != None and label != None:
    # 				splitwords = jieba.cut(comment, cut_all=False)
    # 				for ww in splitwords:
    # 					if ww not in symbols:
    # 						if label == 0:
    # 							f0.write(ww + '\n')
    # 						elif label ==1:
    # 							f1.write(ww + '\n')
    # 	f0.close()
    # 	f1.close()           
                    
    def cmp_data(self):
        Dict = {}
        nokey_0 = 0
        nokey_1 = 0
        key_0 = 0
        key_1 = 0
        file = open('Fws.txt','r')
        
        for i in file.readlines():
            Dict[i] = [0,0] #[pos,neg]

        file2 = open('ntusd-negative.txt','r')
        
        for i in file2.readlines():
            try:
                Dict[i][1] += 1
                key_0 += 1
            except KeyError:
                nokey_0 += 1

        file3 = open('ntusd-positive.txt','r')
        
        for i in file3.readlines():
            try:
                Dict[i][0] += 1
                key_1 += 1
            except KeyError:
                nokey_1 += 1        

        print(nokey_0, nokey_1)
        print(key_0, key_1)
        
        

a = changetype()
# a.changeforme()
# a.changeforyanwen()
a.cmp_data()