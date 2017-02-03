import jieba
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter


class NTUSD_Bayes(object):

	def __init__(self):
		self.trainfile = 'NTUSD_all.xlsx'
		self.testfile = '360test.xlsx'
		self.resultnum = 3
		self.stopwords = []
		self.Dict = {}
		self.P_pos = 0
		self.P_neg = 0
		self.symbols = ('…', ' ', '”', '﹑', '#', '\n', '；', '❤', '〔', '〕', '“', '｀', '•', 'இ', '´', '」', '「', '『', '【', '】', '●', '～', '^', '~', '◆', '.', '@', '』', '', '(',  ')', '"', '\t', '，',  '。', '/',  '！',  '、',  '―',  '？','＠',  '：',   '＃',  '%',   '＆',  '（',  '）',  '《',  '》',  '［',  '］',  '｛',  '｝',  '*',  ',',  '.',   '&',  '!',  '?',  ':',  ';',   '[',  ']', '⋯', '※', '😊', '．', '{',  '}', '　', '..', '-', '+', '=', '_', '<', '>')

	def CreateDict(self):
		f = open('ntusd-negative.txt' , 'r')
		words = f.read().split('\n')
		for w in words:
			self.Dict[w] = [0,0,0,0,0]
							#[Label,posnum,negnum,P(word|pos),P(word|neg)]

		f = open('ntusd-positive.txt' , 'r')
		words = f.read().split('\n')

		for  w in words:
			self.Dict[w] = [1,0,0,0,0]

		f = open('stopwords.txt', 'r')
		self.stopwords = f.read().split('\n')

	# def CreatFeatures(self):
		
		
	# 	jieba.load_userdict("jiebaexpand.txt")

	# 	f = open('neg.txt' , 'w')
	# 	comment = open('negative.txt', 'r').read()
	# 	words = jieba.cut(comment, cut_all=False)
	# 	for word in words:
	# 		if word not in self.symbols:
	# 			f.write(word+',')
	# 	f.close()
		
	# 	#-------------------------------------------------------------------------
		
	# 	f = open('pos.txt' , 'w')
	# 	comment = open('positive.txt', 'r').read()
	# 	words = jieba.cut(comment, cut_all=False)
	# 	for word in words:
	# 		if word not in self.symbols:
	# 			f.write(word+',')
	# 	f.close()

	# 	#--------------------------------------------------------------------

	# 	posnum = 0
	# 	negnum = 0

	# 	f = open('pos.txt' , 'r')
	# 	words = f.read().split(',')
	# 	for w in words:
	# 		try:				
	# 			self.Dict[w][1] += 1
	# 			if self.Dict[w][0] == 0:
	# 				negnum += 1
	# 			elif self.Dict[w][0] == 1:
	# 				posnum += 1
	# 		except KeyError:
	# 			pass

	# 	f.close()

	# 	f = open('neg.txt' , 'r')
	# 	words = f.read().split(',')
	# 	for w in words:
	# 		try:
	# 			self.Dict[w][2] += 1
	# 			if self.Dict[w][0] == 0:
	# 				negnum += 1
	# 			else:
	# 				posnum += 1
	# 		except KeyError:
	# 			pass

	# 	allnum = posnum + negnum

	# 	self.P_pos = posnum / allnum
	# 	self.P_neg = negnum / allnum
	# 	#--------------------------------------------------------------------------

	# 	CountHowManyKeys = 0

	# 	for w in self.Dict.keys():
	# 		if self.Dict[w][1] > 0:
	# 			CountHowManyKeys +=1

	# 	for w in self.Dict.keys():
	# 			self.Dict[w][4] = ( self.Dict[w][2] + 1 ) / ( negnum + CountHowManyKeys ) 
	# 			self.Dict[w][3] = ( self.Dict[w][1] + 1 ) / ( posnum + CountHowManyKeys ) 

	def CreatFeatures(self):

		jieba.load_userdict("jiebaexpand.txt")
		wb1 = load_workbook(self.trainfile)
		sheetnames1 = wb1.get_sheet_names()
		ws1 = wb1.get_sheet_by_name(sheetnames1[0])

		posnum = 0
		negnum = 0

		for i in range(ws1.max_row):
			label = ws1['H{0}'.format(i+1)].value
			comment = ws1['D{0}'.format(i+1)].value
			if comment != None:
				splitwords = jieba.cut(comment, cut_all=False)
				for ww in splitwords:
					if ww not in self.symbols:
						if ww not in self.stopwords:
							try:
								# if label == 0:
								if label == 3:
									self.Dict[ww][2] += 1
									if self.Dict[ww][0] == 0:
										negnum += 1
									else:
										posnum += 1
								# elif label == 1:
								else:
									self.Dict[ww][1] += 1
									if self.Dict[ww][0] == 0:
										negnum += 1
									elif self.Dict[ww][0] == 1:
										posnum += 1
							except KeyError:
								pass

		allnum = posnum + negnum
		self.P_pos = posnum / allnum
		self.P_neg = negnum / allnum
		#--------------------------------------------------------------------------
		CountHowManyKeys = 0

		for w in self.Dict.keys():
			if self.Dict[w][1] > 0:
				CountHowManyKeys +=1

		for w in self.Dict.keys():
				self.Dict[w][4] = ( self.Dict[w][2] + 1 ) / ( negnum + CountHowManyKeys ) 
				self.Dict[w][3] = ( self.Dict[w][1] + 1 ) / ( posnum + CountHowManyKeys ) 

	def Comparsion(self):

		jieba.load_userdict("jiebaexpand.txt")

		wb1 = load_workbook(self.testfile)
		sheetnames1 = wb1.get_sheet_names()
		ws1 = wb1.get_sheet_by_name(sheetnames1[0])

		for i in range(ws1.max_row):
			eMAP_pos = 1
			eMAP_neg = 1
			string = ''
			comment = ws1['D{0}'.format(i+1)].value
			if comment != None:
				splitwords = jieba.cut(comment, cut_all=False)
				for ww in splitwords:
					if ww not in self.symbols:
						if ww not in self.stopwords:
							try:
									# print(self.Dict[ww][4] * self.P_neg,self.Dict[ww][3] * self.P_pos)
									eMAP_neg = eMAP_neg * ( self.Dict[ww][4] * self.P_neg)
									eMAP_pos = eMAP_pos * ( self.Dict[ww][3] * self.P_pos)
									string = string + ww + ','
									ws1['J{0}'.format(i+1)].value = string
							except KeyError:
								pass
				if eMAP_neg > eMAP_pos:
					ws1['I{0}'.format(i+1)].value = 3
				elif eMAP_pos > eMAP_neg:
					ws1['I{0}'.format(i+1)].value = 1
		wb1.save('ANS' + self.testfile)

	def Result(self):
		
		wb1 = load_workbook('ANS' + sel.testfile)
		sheetnames1 = wb1.get_sheet_names()
		ws1 = wb1.get_sheet_by_name(sheetnames1[0])

		original = ''
		predict = ''
		count_ori_1 = 0
		count_ori_0 = 0
		count_pre_0 = 0
		count_pre_1 = 0
		correct_1 = 0
		correct_0 = 0

		for i in range(ws1.max_row):
			# print(type(ws1['I{0}'.format(i+1)]))
			if ws1['I{0}'.format(i+1)].value != None:
				predict = ws1['I{0}'.format(i+1)].value
				if predict == 3:
					count_pre_0 += 1
				elif predict == 1:
					count_pre_1 += 1
			else:
				predict = 5

			if ws1['H{0}'.format(i+1)].value != None:
				original = ws1['H{0}'.format(i+1)].value
				if original == 3 and predict !=5:
					count_ori_0 += 1
				elif (original == 1 or original == 0 or original == 2) and predict !=5:
					count_ori_1 += 1
			else:
				original = 5

			if original == 3 and predict == 3:
				correct_0 += 1
			elif (original == 1 or original == 0 or original == 2) and predict == 1:
				correct_1 += 1
		
			# print(original,predict)
		
		pos_precision = correct_1 / count_pre_1
		neg_precision = correct_0 / count_pre_0
		pos_recall = correct_1 / count_ori_1
		neg_recall =   correct_0 / count_ori_0
		pos_F1 = (2*pos_precision*pos_recall) / (pos_recall + pos_precision)
		neg_F1 = (2*neg_precision*neg_recall) / (neg_recall + neg_precision)
		all_precision = ( correct_1 + correct_0 ) / ( count_pre_0 + count_pre_1 )
		ave_precision = ( pos_precision + neg_precision ) / 2
		ave_recall = ( pos_recall + neg_recall ) / 2
		ave_F1 = ( pos_F1 + neg_F1 ) / 2
		recall = (correct_1 + correct_0) / (count_ori_0 + count_ori_1)
		F1 = (2*all_precision*recall) / (recall + all_precision)

		
		print("pos_precision = {0} / {1} = {2}".format(correct_1,count_pre_1,pos_precision))
		print("pos_recall = {0} / {1} = {2}".format(correct_1,count_ori_1,pos_recall))
		print("pos_F1 = {0} ".format(pos_F1))
		print("neg_precision = {0} / {1} = {2}".format(correct_0,count_pre_0,neg_precision))
		print("neg_recall = {0} / {1} = {2}".format(correct_0,count_ori_0,neg_recall))
		print("neg_F1 = {0} ".format(neg_F1))
		print("ave_precision = {0}".format(ave_precision))
		print("ave_recall = {0}".format(ave_recall))
		print("ave_F1 = {0}".format(ave_F1))
		print("--------------------------------------------------")
		print("precision = {0}".format(all_precision))
		print("recall = {0}".format(recall))
		print("F1 = {0}".format(F1))


		wb = load_workbook('empty.xlsx')
		sheetnames = wb.get_sheet_names()
		ws = wb.get_sheet_by_name(sheetnames[0])

		ws['A1'].value = pos_precision
		ws['A2'].value = pos_recall
		ws['A3'].value = pos_F1
		ws['A4'].value = neg_precision
		ws['A5'].value = neg_recall
		ws['A6'].value = neg_F1
		ws['A7'].value = ave_precision
		ws['A8'].value = ave_recall
		ws['A9'].value = ave_F1
		ws['A11'].value = all_precision
		ws['A12'].value = recall
		ws['A13'].value = F1

		wb.save('result_NB3_{0}.xlsx'.format(self.resultnum))


a = NTUSD_Bayes()
# a.CreateDict()
# a.CreatFeatures()
# a.Comparsion()
a.Result()




