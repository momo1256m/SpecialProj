import requests, json, xlwt
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

url = 'https://graph.facebook.com/'
token = '1214654765221490|PMLcik1KIIoJF4--xv45P12kgyA'
# token = 'EAACEdEose0cBAMxRkN0PGR0JYdq2R3Fr0v9qeGOZCIYR1qFHovKPRkmexFb9AoJQy7tjIPKbuIZBbPdEmzAJWr7cjQulITa24UgheDnlJTvZAwVUIQTGJRl0W0opNbexTmJHI8NFngh9eoVbshIZB7fdDPzsYJnAkZAhZCE7HuZAgZDZD'

end = '2015-09-15'
start = '2015-09-05'

class crawlerforFB(object):

	
	def GetSpecifyPostID(self):
		res = requests.get(url + 'v2.8/MaYingjeou?access_token={0}&fields=posts.limit(100).until({1}).since({2})'.format(token,end,start))
		jsondata = json.loads(res.text , encoding="utf-8")
		f = open('PostsID.json' , 'w' , encoding= 'utf-8')
		time = []
		j = 0
		escape = 0

		for i in jsondata['posts']['data']:
			f.write(i['id'] + '\n')
			# time.append(i['created_time'][:4]+i['created_time'][5:7]+i['created_time'][8:10])
			# if(int(time[j]) >= start and int(time[j]) <= end):
			# 	f.write(i['id']+'\n')
			# j = j+1
		
		re = requests.get(jsondata['posts']['paging']['next'])
		if 'next' in jsondata['posts']['paging']:
			for i in jsondata['posts']['data']:
				f.write(i['id']+'\n')				
				# time.append(i['created_time'][:4]+i['created_time'][5:7]+i['created_time'][8:10])
				# if(int(time[j]) >= start and int(time[j]) <= end):
				# j = j+1
			if 'next' in jsondata['posts']['paging']:
				re = requests.get(jsondata['posts']['paging']['next'])
				jsondata = json.loads(re.text)
		try:
			while 'previous' in jsondata['paging']:
				for i in jsondata['data']:
					f.write(i['id']+'\n')
					# time.append(i['created_time'][:4]+i['created_time'][5:7]+i['created_time'][8:10])
					# if(int(time[j]) >= start and int(time[j]) <= end):
				if 'next' in jsondata['paging']:
					re = requests.get(jsondata['paging']['next'])
					jsondata = json.loads(re.text)
				else:
					break
		except KeyError:
			pass
			# 	elif(int(time[j]) < start):
			# 		escape = 1
			# 		break
			# 	j = j+1
			# if(escape == 1):
			# 	break
			# else:	
			# 	if 'next' in jsondata['paging']:
			# 		re = requests.get(jsondata['paging']['next'])
			# 		jsondata = json.loads(re.text)
			# 	else:
			# 		break

		f.close()

	def ParseComment(self):
	 	f = open('PostsID.json' , 'r' , encoding= 'utf-8')
	 	index = 1
	 	for line in f.readlines():
	 		res = requests.get(url + 'v2.3/{0}?access_token={1}&fields=comments.limit(100)'.format(line,token))
	 		jsonObj = json.loads(res.text)
	 		# with open('data.json' , 'w' , encoding= 'utf-8') as f:
	 		# 	json.dump(jsonObj, f)
	 		wbk = xlwt.Workbook()
	 		sheet = wbk.add_sheet('sheet 1')			
	 		x=0
	 		post_id= 0
	 		ID=1
	 		name = 2
	 		message = 3
	 		message_id = 4
	 		like_count = 5
	 		created_time = 6
	 		print(line)
	 		for i in jsonObj['comments']['data']:
	 			sheet.write(x,post_id,line)
	 			sheet.write(x,ID,i['from']['id'])
	 			sheet.write(x,name,i['from']['name'])
	 			sheet.write(x,message,i['message'])
	 			sheet.write(x,message_id,i['id'])
	 			# sheet.write(x,like_count,i['like_count'])
	 			sheet.write(x,created_time,i['created_time'])
	 			x=x+1
	 		try:
	 			re = requests.get(jsonObj['comments']['paging']['next'])
	 			jsonObj = json.loads(re.text)
	 		except KeyError:
	 			pass
	 		while 'previous' in jsonObj:
	 			for i in jsonObj['data']:
	 				sheet.write(x,ID,line)
	 				sheet.write(x,name,i['from']['name'])
	 				sheet.write(x,message,i['message'])
	 				sheet.write(x,message_id,i['id'])
	 				# sheet.write(x,like_count,i['like_count'])
	 				sheet.write(x,created_time,i['created_time'])
	 				x=x+1
	 				if 'next' in jsonObj['paging']:
	 					re = requests.get(jsonObj['paging']['next'])
	 					jsonObj = json.loads(re.text)
	 				else:
	 					break

	 		wbk.save('{0}.xlsx'.format(index))
	 		index += 1
	 	
	 	f.close()


	# def ParseComment(self):
	#  	f = open('PostsID.json' , 'r' , encoding= 'utf-8')
	#  	wb1 = load_workbook('empty.xlsx')
	#  	sheetnames1 = wb1.get_sheet_names()
	#  	ws1 = wb1.get_sheet_by_name(sheetnames1[0])
	 	
	#  	index = 1
	#  	count = 1

	#  	for line in f.readlines():
	#  		res = requests.get(url + 'v2.3/{0}?access_token={1}&fields=comments.limit(100)'.format(line,token))
	#  		jsonObj = json.loads(res.text)
	#  		for i in jsonObj['comments']['data']:
	#  			i_count = str(count)
	#  			ws1['A' + i_count].value = line
	#  			ws1['B' + i_count].value = i['from']['id']
	#  			ws1['C' + i_count].value = i['from']['name']
	#  			ws1['D' + i_count].value = i['message']
	#  			ws1['E' + i_count].value = i['id']
	#  			# ws1['F'][count].value = i['like_count']
	#  			ws1['G' + i_count].value = i['created_time']
	#  			count += 1
	#  		try:
	#  			re = requests.get(jsonObj['comments']['paging']['next'])
	#  			jsonObj = json.loads(re.text)
	#  		except KeyError:
	#  			pass
	#  		while 'previous' in jsonObj:
	#  			for i in jsonObj['data']:
	#  				i_count = str(count)
	#  				ws1['A' + i_count].value = line
	# 	 			ws1['B' + i_count].value = i['from']['id']
	# 	 			ws1['C' + i_count].value = i['from']['name']
	# 	 			ws1['D' + i_count].value = i['message']
	# 	 			ws1['E' + i_count].value = i['id']
	# 	 			# ws1['F'][count].value = i['like_count']
	# 	 			ws1['G' + i_count].value = i['created_time']
	# 	 			count += 1
	#  				if 'next' in jsonObj['paging']:
	#  					re = requests.get(jsonObj['paging']['next'])
	#  					jsonObj = json.loads(re.text)
	#  				else:
	#  					break
	#  	wb1.save('{0}.xlsx'.format(index))
	#  	index += 1

crawler = crawlerforFB()
crawler.GetSpecifyPostID()
crawler.ParseComment()	