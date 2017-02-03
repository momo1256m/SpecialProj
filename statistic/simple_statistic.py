import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.utils import get_column_letter


class simple_statistic(object):
	
	def statistic_for_one(self):

		data = {}
		datasorted = []
		items = []
		dataforstore = []

		wb1 = load_workbook('4.xlsx')
		sheetnames1 = wb1.get_sheet_names()
		ws1 = wb1.get_sheet_by_name(sheetnames1[0])

		for j in range(700):
			ID = ws1['B{0}'.format(j+1)].value
			label = ws1['H{0}'.format(j+1)].value
			name = ws1['C{0}'.format(j+1)].value
			if ID != None and label != None:
				try:
					if label == 0:
						data[ID][1] += 1
						data[ID][3] += 1
					elif label == 1:
						data[ID][1] += 1
						data[ID][2] += 1
				except KeyError:
					data[ID] = [name,0,0,0]
					if label == 0:
						data[ID][1] += 1
						data[ID][3] += 1
					elif label == 1:
						data[ID][1] += 1
						data[ID][2] += 1

		items = data.items()
		datasorted = [[v[1][1],v[0]] for v in items]
		datasorted.sort()

		wb2 = load_workbook('4_analysis.xlsx')
		sheetnames2 = wb2.get_sheet_names()
		ws2 = wb2.get_sheet_by_name(sheetnames2[0])

		ws2['A1'].value = 'ID'
		ws2['B1'].value = 'Name'
		ws2['C1'].value = 'All_Comment'
		ws2['D1'].value = 'Pos_Comment'
		ws2['E1'].value = 'Neg_Comment'

		for k in range(0,len(datasorted)):
			position = k+2
			ws2['A{0}'.format(position)].value = datasorted[k][1]
			ws2['B{0}'.format(position)].value = data[datasorted[k][1]][0]
			ws2['C{0}'.format(position)].value = data[datasorted[k][1]][1]
			ws2['D{0}'.format(position)].value = data[datasorted[k][1]][2]
			ws2['E{0}'.format(position)].value = data[datasorted[k][1]][3]

		wb2.save('4_analysis.xlsx')
			# print(dataforstore)	
				
	def statistic_for_all(self):

		data = {}
		datasorted = []
		items = []
		dataforstore = []

		for i in range(0,4):
			wb1 = load_workbook('{0}.xlsx'.format(i+1))
			sheetnames1 = wb1.get_sheet_names()
			ws1 = wb1.get_sheet_by_name(sheetnames1[0])

			for j in range(700):
				ID = ws1['B{0}'.format(j+1)].value
				label = ws1['H{0}'.format(j+1)].value
				name = ws1['C{0}'.format(j+1)].value
				if ID != None and label != None:
					if i == 0:
						try:
							if label == 0:
								data[ID][1] += 1
								data[ID][2] += 1
								data[ID][4] += 1
							elif label == 1:
								data[ID][1] += 1
								data[ID][2] += 1
								data[ID][3] += 1
						except KeyError:
							data[ID] = [name,0,0,0,0,0,0,0,0,0,0,0,0,0]
							#[name,allsum,1_sum,1_pos,1_neg,...]
							if label == 0:
								data[ID][1] += 1
								data[ID][2] += 1
								data[ID][4] += 1
							elif label == 1:
								data[ID][1] += 1
								data[ID][2] += 1
								data[ID][3] += 1
					if i == 1:
						try:
							if label == 0:
								data[ID][1] += 1
								data[ID][5] += 1
								data[ID][7] += 1
							elif label == 1:
								data[ID][1] += 1
								data[ID][5] += 1
								data[ID][6] += 1
						except KeyError:
							data[ID] = [name,0,0,0,0,0,0,0,0,0,0,0,0,0]
							#[name,allsum,1_sum,1_pos,1_neg,...]
							if label == 0:
								data[ID][1] += 1
								data[ID][5] += 1
								data[ID][7] += 1
							elif label == 1:
								data[ID][1] += 1
								data[ID][5] += 1
								data[ID][6] += 1
					if i == 2:
						try:
							if label == 0:
								data[ID][1] += 1
								data[ID][8] += 1
								data[ID][10] += 1
							elif label == 1:
								data[ID][1] += 1
								data[ID][8] += 1
								data[ID][9] += 1
						except KeyError:
							data[ID] = [name,0,0,0,0,0,0,0,0,0,0,0,0,0]
							#[name,allsum,1_sum,1_pos,1_neg,...]
							if label == 0:
								data[ID][1] += 1
								data[ID][8] += 1
								data[ID][10] += 1
							elif label == 1:
								data[ID][1] += 1
								data[ID][8] += 1
								data[ID][9] += 1
					if i == 3:
						try:
							if label == 0:
								data[ID][1] += 1
								data[ID][11] += 1
								data[ID][13] += 1
							elif label == 1:
								data[ID][1] += 1
								data[ID][11] += 1
								data[ID][12] += 1
						except KeyError:
							data[ID] = [name,0,0,0,0,0,0,0,0,0,0,0,0,0]
							#[name,allsum,1_sum,1_pos,1_neg,...]
							if label == 0:
								data[ID][1] += 1
								data[ID][11] += 1
								data[ID][13] += 1
							elif label == 1:
								data[ID][1] += 1
								data[ID][11] += 1
								data[ID][12] += 1

		
		items = data.items()
		datasorted = [[v[1][1],v[0]] for v in items]
		datasorted.sort()

		wb2 = load_workbook('all.xlsx')
		sheetnames2 = wb2.get_sheet_names()
		ws2 = wb2.get_sheet_by_name(sheetnames2[0])

		ws2['A1'].value = 'ID'
		ws2['B1'].value = 'Name'
		ws2['C1'].value = 'All_Comment'
		ws2['D1'].value = '1_all_Comment'
		ws2['E1'].value = '1_Pos_Comment'
		ws2['F1'].value = '1_Neg_Comment'
		ws2['G1'].value = '2_all_Comment'
		ws2['H1'].value = '2_Pos_Comment'
		ws2['I1'].value = '2_Neg_Comment'
		ws2['J1'].value = '3_all_Comment'
		ws2['K1'].value = '3_Pos_Comment'
		ws2['L1'].value = '3_Neg_Comment'
		ws2['M1'].value = '4_all_Comment'
		ws2['N1'].value = '4_Pos_Comment'
		ws2['O1'].value = '4_Neg_Comment'

		for k in range(0,len(datasorted)):
			position = k+2
			ws2['A{0}'.format(position)].value = datasorted[k][1]
			ws2['B{0}'.format(position)].value = data[datasorted[k][1]][0]
			ws2['C{0}'.format(position)].value = data[datasorted[k][1]][1]
			ws2['D{0}'.format(position)].value = data[datasorted[k][1]][2]
			ws2['E{0}'.format(position)].value = data[datasorted[k][1]][3]
			ws2['F{0}'.format(position)].value = data[datasorted[k][1]][4]
			ws2['G{0}'.format(position)].value = data[datasorted[k][1]][5]
			ws2['H{0}'.format(position)].value = data[datasorted[k][1]][6]
			ws2['I{0}'.format(position)].value = data[datasorted[k][1]][7]
			ws2['J{0}'.format(position)].value = data[datasorted[k][1]][8]
			ws2['K{0}'.format(position)].value = data[datasorted[k][1]][9]
			ws2['L{0}'.format(position)].value = data[datasorted[k][1]][10]
			ws2['M{0}'.format(position)].value = data[datasorted[k][1]][11]
			ws2['N{0}'.format(position)].value = data[datasorted[k][1]][12]
			ws2['O{0}'.format(position)].value = data[datasorted[k][1]][13]


		wb2.save('all.xlsx')	
			# print(dataforstore)	

			

	def all_pos(self):
		
		wb2 = load_workbook('all_pos.xlsx')
		sheetnames2 = wb2.get_sheet_names()
		ws2 = wb2.get_sheet_by_name(sheetnames2[0])

		position = 2

		ws2['A1'].value = 'ID'
		ws2['B1'].value = 'Name'
		ws2['C1'].value = 'All_Comment'
		ws2['D1'].value = '1_all_Comment'
		ws2['E1'].value = '1_Pos_Comment'
		ws2['F1'].value = '1_Neg_Comment'
		ws2['G1'].value = '2_all_Comment'
		ws2['H1'].value = '2_Pos_Comment'
		ws2['I1'].value = '2_Neg_Comment'
		ws2['J1'].value = '3_all_Comment'
		ws2['K1'].value = '3_Pos_Comment'
		ws2['L1'].value = '3_Neg_Comment'
		ws2['M1'].value = '4_all_Comment'
		ws2['N1'].value = '4_Pos_Comment'
		ws2['O1'].value = '4_Neg_Comment'	

		wb1 = load_workbook('all.xlsx')
		sheetnames1 = wb1.get_sheet_names()
		ws1 = wb1.get_sheet_by_name(sheetnames1[0])

		for j in range(1308):
			if ws1['A{0}'.format(j+2)].value != None:
				_1_Pos_Comment = ws1['E{0}'.format(j+2)].value
				_1_Neg_Comment = ws1['F{0}'.format(j+2)].value
				_2_Pos_Comment = ws1['H{0}'.format(j+2)].value
				_2_Neg_Comment = ws1['I{0}'.format(j+2)].value
				_3_Pos_Comment = ws1['K{0}'.format(j+2)].value
				_3_Neg_Comment = ws1['L{0}'.format(j+2)].value
				_4_Pos_Comment = ws1['N{0}'.format(j+2)].value
				_4_Neg_Comment = ws1['O{0}'.format(j+2)].value
			
				if _1_Neg_Comment == 0 and _2_Neg_Comment == 0 and _3_Neg_Comment == 0 and _4_Neg_Comment == 0:
					ws2['A{0}'.format(position)].value = ws1['A{0}'.format(j+2)].value
					ws2['B{0}'.format(position)].value = ws1['B{0}'.format(j+2)].value
					ws2['C{0}'.format(position)].value = ws1['C{0}'.format(j+2)].value
					ws2['D{0}'.format(position)].value = ws1['D{0}'.format(j+2)].value
					ws2['E{0}'.format(position)].value = ws1['E{0}'.format(j+2)].value
					ws2['F{0}'.format(position)].value = ws1['F{0}'.format(j+2)].value
					ws2['G{0}'.format(position)].value = ws1['G{0}'.format(j+2)].value
					ws2['H{0}'.format(position)].value = ws1['H{0}'.format(j+2)].value
					ws2['I{0}'.format(position)].value = ws1['I{0}'.format(j+2)].value
					ws2['J{0}'.format(position)].value = ws1['J{0}'.format(j+2)].value
					ws2['K{0}'.format(position)].value = ws1['K{0}'.format(j+2)].value
					ws2['L{0}'.format(position)].value = ws1['L{0}'.format(j+2)].value
					ws2['M{0}'.format(position)].value = ws1['M{0}'.format(j+2)].value
					ws2['N{0}'.format(position)].value = ws1['N{0}'.format(j+2)].value
					ws2['O{0}'.format(position)].value = ws1['O{0}'.format(j+2)].value
					position = position + 1
		wb2.save('all_pos.xlsx')

	def all_neg(self):
		wb2 = load_workbook('all_neg.xlsx')
		sheetnames2 = wb2.get_sheet_names()
		ws2 = wb2.get_sheet_by_name(sheetnames2[0])

		position = 2

		ws2['A1'].value = 'ID'
		ws2['B1'].value = 'Name'
		ws2['C1'].value = 'All_Comment'
		ws2['D1'].value = '1_all_Comment'
		ws2['E1'].value = '1_Pos_Comment'
		ws2['F1'].value = '1_Neg_Comment'
		ws2['G1'].value = '2_all_Comment'
		ws2['H1'].value = '2_Pos_Comment'
		ws2['I1'].value = '2_Neg_Comment'
		ws2['J1'].value = '3_all_Comment'
		ws2['K1'].value = '3_Pos_Comment'
		ws2['L1'].value = '3_Neg_Comment'
		ws2['M1'].value = '4_all_Comment'
		ws2['N1'].value = '4_Pos_Comment'
		ws2['O1'].value = '4_Neg_Comment'	

		wb1 = load_workbook('all.xlsx')
		sheetnames1 = wb1.get_sheet_names()
		ws1 = wb1.get_sheet_by_name(sheetnames1[0])

		for j in range(1308):
			if ws1['A{0}'.format(j+2)].value != None:
				_1_Pos_Comment = ws1['E{0}'.format(j+2)].value
				_1_Neg_Comment = ws1['F{0}'.format(j+2)].value
				_2_Pos_Comment = ws1['H{0}'.format(j+2)].value
				_2_Neg_Comment = ws1['I{0}'.format(j+2)].value
				_3_Pos_Comment = ws1['K{0}'.format(j+2)].value
				_3_Neg_Comment = ws1['L{0}'.format(j+2)].value
				_4_Pos_Comment = ws1['N{0}'.format(j+2)].value
				_4_Neg_Comment = ws1['O{0}'.format(j+2)].value
			
				if _1_Pos_Comment == 0 and _2_Pos_Comment == 0 and _3_Pos_Comment == 0 and _4_Pos_Comment == 0:
					ws2['A{0}'.format(position)].value = ws1['A{0}'.format(j+2)].value
					ws2['B{0}'.format(position)].value = ws1['B{0}'.format(j+2)].value
					ws2['C{0}'.format(position)].value = ws1['C{0}'.format(j+2)].value
					ws2['D{0}'.format(position)].value = ws1['D{0}'.format(j+2)].value
					ws2['E{0}'.format(position)].value = ws1['E{0}'.format(j+2)].value
					ws2['F{0}'.format(position)].value = ws1['F{0}'.format(j+2)].value
					ws2['G{0}'.format(position)].value = ws1['G{0}'.format(j+2)].value
					ws2['H{0}'.format(position)].value = ws1['H{0}'.format(j+2)].value
					ws2['I{0}'.format(position)].value = ws1['I{0}'.format(j+2)].value
					ws2['J{0}'.format(position)].value = ws1['J{0}'.format(j+2)].value
					ws2['K{0}'.format(position)].value = ws1['K{0}'.format(j+2)].value
					ws2['L{0}'.format(position)].value = ws1['L{0}'.format(j+2)].value
					ws2['M{0}'.format(position)].value = ws1['M{0}'.format(j+2)].value
					ws2['N{0}'.format(position)].value = ws1['N{0}'.format(j+2)].value
					ws2['O{0}'.format(position)].value = ws1['O{0}'.format(j+2)].value
					position = position + 1
		wb2.save('all_neg.xlsx')



a = simple_statistic()
# a.statistic_for_one()
# a.statistic_for_all()
a.all_pos()
# a.all_neg()

